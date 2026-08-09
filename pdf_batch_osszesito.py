import os
import re
import pickle
import csv
from pypdf import PdfReader
from pathlib import Path

from openpyxl import Workbook

from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow


# =========================
# BEÁLLÍTÁSOK
# =========================
INPUT_DIR = "input"

# Ide csak a hibátlan Excel fájlok kerülnek
OUTPUT_DIR = "output"

# Ide kerülnek a hibás Excel fájlok, hogy külön át tudd nézni
BAD_OUTPUT_DIR = "hibas_output"

# Ide csak a hibás fájlok TXT debug fájljai kerülnek
TXT_DEBUG_DIR = "txt_debug"

CHECK_FILE = "ellenorzes.xlsx"
SUMMARY_FILE = "osszesitett_tetelek.xlsx"

TOKEN_FILE = "token.json"
CLIENT_SECRET_FILE = "client_secret.json"

SCOPES = ["https://www.googleapis.com/auth/drive"]

KULCS_OUTPUT_FILE = "kulcs_bejovo_szamlak.csv"

# Demó / ügyfélspecifikus Kulcs-Ügyvitel törzsadatok
KULCS_PARTNER_KOD = "MPX001"
KULCS_PARTNER_NEV = "MOTORPART X POLSKA Sp. z o.o."
KULCS_IRANYITOSZAM = "40-001"
KULCS_VAROS = "Katowice"
KULCS_UTCA = "Przemysłowa"
KULCS_KOZTERULET = "ul."
KULCS_HAZSZAM = "18"
KULCS_FIZETESI_MOD = "Banki átutalás"
KULCS_AFA_KULCS = "27%"
KULCS_MENNYISEGI_EGYSEG = "db"
KULCS_PENZNEM = "HUF"
KULCS_ARFOLYAM = 1
# =========================


# -------------------------
# Google hitelesítés
# -------------------------
def authenticate():
    creds = None

    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, "rb") as token:
            creds = pickle.load(token)

    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
        with open(TOKEN_FILE, "wb") as token:
            pickle.dump(creds, token)

    if not creds or not creds.valid:
        flow = InstalledAppFlow.from_client_secrets_file(
            CLIENT_SECRET_FILE,
            SCOPES
        )
        creds = flow.run_local_server(port=0)

        with open(TOKEN_FILE, "wb") as token:
            pickle.dump(creds, token)

    return build("drive", "v3", credentials=creds)


# -------------------------
# PDF -> Google Docs
# -------------------------
def pdf_to_google_doc(service, pdf_path, doc_name):
    file_metadata = {
        "name": doc_name,
        "mimeType": "application/vnd.google-apps.document"
    }

    media = MediaFileUpload(pdf_path, mimetype="application/pdf")

    file = service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id"
    ).execute()

    return file.get("id")


# -------------------------
# Google Docs -> TXT
# -------------------------
def get_doc_text(service, file_id):
    export = service.files().export(
        fileId=file_id,
        mimeType="text/plain"
    ).execute()

    return export.decode("utf-8")


# -------------------------
# Drive fájl törlés
# -------------------------
def delete_file(service, file_id):
    try:
        service.files().delete(fileId=file_id).execute()
    except Exception:
        pass

# -------------------------
# PDF -> TXT közvetlenül
# -------------------------

def get_pdf_text(pdf_path):
    reader = PdfReader(pdf_path)

    text_parts = []

    for page in reader.pages:
        text = page.extract_text()

        if text:
            text_parts.append(text)

    return "\n".join(text_parts)

# -------------------------
# Alap segédek
# -------------------------
def hu_to_float(value):
    if value is None:
        return None

    s = str(value).strip()
    if not s:
        return None

    s = s.replace(" ", "").replace(",", ".")

    try:
        return float(s)
    except ValueError:
        return None


def normalize_text(text: str) -> str:
    text = text.replace("\r", "\n")

    # szétesett fejléc javítása
    text = text.replace("Egysé\ng", "Egység")
    text = text.replace("Mennyisé\ng", "Mennyiség")
    text = text.replace("A szállító  országa", "A szállító országa")
    text = text.replace("A szállító \nországa", "A szállító országa")
    text = text.replace("Bruttó  súly", "Bruttó súly")
    text = text.replace("Bruttó \nsúly", "Bruttó súly")
    text = text.replace("Bruttó  tömeg", "Bruttó tömeg")
    text = text.replace("Bruttó \ntömeg", "Bruttó tömeg")

    # GMBH és ár összeragadás javítása:
    # GMBH6 290,00 -> GMBH 6 290,00
    # GMBH7 254,00 -> GMBH 7 254,00
    text = re.sub(r"(GMBH)(?=\d)", r"\1 ", text)

    # MOTO. és ár összeragadás javítása:
    # MOTO.1 135,00 -> MOTO. 1 135,00
    text = re.sub(r"(MOTO\.)(?=\d)", r"\1 ", text)

    # MGZ cikkszám összeragadás javítása:
    # KERÉKFELFÜGGESZTÉSMGZ-116014 -> KERÉKFELFÜGGESZTÉS MGZ-116014
    text = re.sub(r"([A-ZÁÉÍÓÖŐÚÜŰ0-9])(?=MGZ-\d+)", r"\1 ", text)

    # MGA cikkszám összeragadás javítása:
    # 206MGA-5718 -> 206 MGA-5718
    text = re.sub(r"([A-ZÁÉÍÓÖŐÚÜŰ0-9])(?=MGA-\d+)", r"\1 ", text)

    # VAL cikkszám összeragadás javítása:
    # 1.6MJVAL574645 -> 1.6MJ VAL574645
    text = re.sub(r"([A-ZÁÉÍÓÖŐÚÜŰ0-9])(?=VAL\d+)", r"\1 ", text)

    # összeragadt szt javítása:
    # OSR62210CBB-2HBszt -> OSR62210CBB-2HB szt
    text = re.sub(r"(?i)([A-Z0-9\-.\/])(?=(szt|SZT|kpl|KPL)\b)", r"\1 ", text)

    fixed_lines = []
    for line in text.split("\n"):
        line = re.sub(r"[ \t]+", " ", line).strip()
        fixed_lines.append(line)

    return "\n".join(fixed_lines)


def save_txt_debug(pdf_name: str, text: str):
    os.makedirs(TXT_DEBUG_DIR, exist_ok=True)

    txt_name = Path(pdf_name).stem + ".txt"
    txt_path = os.path.join(TXT_DEBUG_DIR, txt_name)

    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(text)


def parse_invoice_no(text):
    m = re.search(
        r"\b[A-Z0-9]+/\d{4}/\d{2}/\d{3}\b",
        text
    )

    return m.group(0) if m else ""

def parse_invoice_date(text):
    m = re.search(
        r"Kiállítás:\s*(\d{4}-\d{2}-\d{2})",
        text
    )

    if m:
        return m.group(1)

    # tartalék megoldás
    m = re.search(
        r"\b(\d{4}-\d{2}-\d{2})\b",
        text
    )

    return m.group(1) if m else ""

def parse_due_date(text):
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    for i, line in enumerate(lines):
        if line.startswith("Fizetési határidő"):
            dates = []

            for candidate in lines[i + 1:i + 7]:
                if re.fullmatch(r"\d{4}-\d{2}-\d{2}", candidate):
                    dates.append(candidate)

            # A blokkban az első dátum a számla dátuma,
            # a második a fizetési határidő.
            if len(dates) >= 2:
                return dates[1]

            if dates:
                return dates[-1]

    return ""

def parse_total_row(text: str):
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    for i, line in enumerate(lines):

        if line == "Teljes" and i + 4 < len(lines):

            if lines[i + 3] == "Bruttó tömeg":

                try:
                    total_qty = hu_to_float(lines[i + 1])
                    total_amount = hu_to_float(lines[i + 2])
                    total_brutto = hu_to_float(lines[i + 4])

                    return total_qty, total_amount, total_brutto

                except Exception:
                    pass

    return None, None, None

def close_enough(a, b, tolerance=0.02):
    if a is None or b is None:
        return False

    return abs(a - b) <= tolerance


# -------------------------
# Parsing
# -------------------------
def split_name_and_code(prefix: str):
    """
    Az SZT előtti részből kinyeri a terméknevet és cikkszámot.

    Alapeset:
    - az utolsó token a cikkszám, ha tartalmaz számot.

    Speciális esetek:
    - QBWS 0366 A
    - FOLK VANILLA
    - FOLK GREEN TEA
    """
    prefix = prefix.strip()
    if not prefix:
        return "", ""

    parts = prefix.split()
    if not parts:
        return "", ""

    # Speciális cikkszám: QBWS 0366 A
    if len(parts) >= 3:
        t1 = parts[-3]
        t2 = parts[-2]
        t3 = parts[-1]

        if (
            re.fullmatch(r"[A-Z]{2,}", t1)
            and re.fullmatch(r"\d+", t2)
            and re.fullmatch(r"[A-Z]", t3)
        ):
            name = " ".join(parts[:-3]).strip()
            code = f"{t1} {t2} {t3}"
            return name, code

    # Speciális cikkszám: FOLK GREEN TEA
    if len(parts) >= 3:
        if parts[-3] == "FOLK":
            name = " ".join(parts[:-3]).strip()
            code = " ".join(parts[-3:]).strip()
            return name, code

    # Speciális cikkszám: FOLK VANILLA
    if len(parts) >= 2:
        if parts[-2] == "FOLK":
            name = " ".join(parts[:-2]).strip()
            code = " ".join(parts[-2:]).strip()
            return name, code

        # Speciális cikkszám: SMELLY OWL RED FRUIT
    if len(parts) >= 4:
        if parts[-4:] == ["SMELLY", "OWL", "RED", "FRUIT"]:
            name = " ".join(parts[:-4]).strip()
            code = " ".join(parts[-4:]).strip()
            return name, code
        
    # Normál eset: az utolsó token a cikkszám
    last_token = parts[-1]

    if any(ch.isdigit() for ch in last_token):
        name = " ".join(parts[:-1]).strip()
        return name, last_token

    return prefix, ""

def prepare_body(text: str):
    """
    A fejléc utáni, Teljes előtti részt adja vissza.
    A Választék blokkokat levágja.
    Az oldalfejléceket és egyéb zajsorokat kiszűri.
    """
    text = normalize_text(text)

    # Az első tényleges táblázatfejléc után indulunk
    m = re.search(
        r"Árú megnevezése.*?Bruttó tömeg",
        text,
        flags=re.DOTALL
    )
    if m:
        text = text[m.end():]

    # Az első Teljes sorig nézzük
    end_idx = text.find("Teljes")
    if end_idx != -1:
        text = text[:end_idx]

    cleaned_lines = []

    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue

        # Oldalfejléc / zaj kiszűrése
        if line.startswith("Auto Partner SA"):
            continue
        if line.startswith("ul. Ekonomiczna"):
            continue
        if line.startswith("43-150"):
            continue
        if line == "Polska":
            continue
        if line.startswith("Számlálási és mérlegelési protokoll"):
            continue
        if line.startswith("ÁFA számla:"):
            continue
        if line.startswith("Számla dátuma:"):
            continue
        if line.startswith("Címzett:"):
            continue
        if line.startswith("Oldal"):
            continue
        if line.startswith("Csomagok"):
            continue
        if line.startswith("Mennyiség"):
            continue
        if line.startswith("Nettó tömeg:"):
            continue
        if line.startswith("Bruttó súly:"):
            continue
        if line.startswith("Csomagolás súlya:"):
            continue

        # A választék részt levágjuk a sor végéről
        line = re.sub(r"\s+\d+\s+Választék:.*$", "", line).strip()

        if not line:
            continue

        cleaned_lines.append(line)

    return " ".join(cleaned_lines)


def parse_text(text: str):
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    results = []
    seen = set()

    units = {"SZT", "KPL"}

    number_pattern = re.compile(r"^-?\d+,\d{2}$")
    amount_pattern = re.compile(r"^-?[\d ]+,\d{2}$")

    for i, line in enumerate(lines):

        if line.upper() not in units:
            continue

        # Kell 2 sor előtte és 7 sor utána
        if i < 2 or i + 7 >= len(lines):
            continue

        termeknev = lines[i - 2]
        cikkszam = lines[i - 1]

        mennyiseg = lines[i + 1]
        orszag = lines[i + 2]
        gyarto = lines[i + 3]
        netto_ar = lines[i + 4].replace(" ", "")
        valuta = lines[i + 5]
        brutto_suly = lines[i + 6]
        brutto_tomeg = lines[i + 7]

        # Ellenőrzések, hogy tényleg tételsort találtunk
        if not number_pattern.match(mennyiseg):
            continue

        if not re.fullmatch(r"[A-Z]{2}", orszag):
            continue

        if not amount_pattern.match(lines[i + 4]):
            continue

        if not re.fullmatch(r"[A-Z]{3}", valuta):
            continue

        if not number_pattern.match(brutto_suly):
            continue

        if not number_pattern.match(brutto_tomeg):
            continue

        # 0 értékű korrekciós sorok kihagyása
        if netto_ar in ("0,00", "-0,00"):
            continue

        row = (
            termeknev,
            cikkszam,
            mennyiseg,
            orszag,
            gyarto,
            netto_ar,
            valuta,
            brutto_suly,
            brutto_tomeg
        )

        if row in seen:
            continue

        seen.add(row)
        results.append(row)

    return results

# -------------------------
# Excel mentés
# -------------------------
def save_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Adatok"

    ws.append([
        "Forras_fajl",
        "Szamla_szam",
        "Szamla_datum",
        "Termeknev",
        "Cikkszam",
        "Mennyiseg",
        "Szallito_orszaga",
        "Gyarto",
        "Mennyiseg_ar",
        "Valuta",
        "Brutto_suly",
        "Brutto_tomeg",
    ])

    for row in rows:
        ws.append(row)

    wb.save(output_path)


def save_check_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ellenorzes"

    ws.append([
        "Forras_fajl",
        "Szamla_szam",
        "Szamla_datum",
        "Sorok_szama",
        "PDF_darabszam",
        "Kinyert_darabszam",
        "Darabszam_egyezik",
        "PDF_osszeg",
        "Kinyert_osszeg",
        "Osszeg_egyezik",
        "PDF_brutto_tomeg",
        "Kinyert_brutto_tomeg",
        "Brutto_tomeg_egyezik",
        "Statusz",
        "Megjegyzes",
    ])

    for row in rows:
        ws.append(row)

    wb.save(output_path)

def save_summary_excel(rows, output_path):
    grouped = {}

    for row in rows:
        termeknev = row[3]
        cikkszam = row[4]
        mennyiseg = hu_to_float(row[5]) or 0

        key = cikkszam

        if key not in grouped:
            grouped[key] = {
                "termeknev": termeknev,
                "mennyiseg": 0,
            }

        grouped[key]["mennyiseg"] += mennyiseg

    wb = Workbook()
    ws = wb.active
    ws.title = "Osszesitett tetelek"

    ws.append([
        "Cikkszam",
        "Termeknev",
        "Osszes_mennyiseg",
    ])

    for cikkszam, adat in sorted(grouped.items()):
        if adat["mennyiseg"] > 0:
            ws.append([
                cikkszam,
                adat["termeknev"],
                adat["mennyiseg"],
            ])

    wb.save(output_path)

def kulcs_decimal(value):
    if value is None:
        return ""

    return f"{value:.4f}".rstrip("0").rstrip(".").replace(".", ",")


def kulcs_date(value):
    if not value:
        return ""

    # 2026-07-01 -> 2026.07.01
    return value.replace("-", ".")


def save_kulcs_csv(rows, output_path):
    headers = [
        "Sorszám",
        "Pénznem",
        "Árfolyam",
        "Ügyfél kód",
        "Ügyfél Név",
        "Irányítószám",
        "Város",
        "Utca",
        "Adószám",
        "Fizetési mód",
        "Kelt",
        "Teljesítés",
        "Esedékesség",
        "Terméknév",
        "Mennyiség",
        "Mennyiségi egység",
        "Egységár",
        "Megjegyzés",
        "Áfa kulcs",
        "Munkaszám",
        "Részlegszám",
        "Számla megjegyzés",
        "TERMEKKOD",
        "Bizonylatszám",
        "Elszámoló időszak  záró dátuma",
        "Áfa árfolyam",
        "Termék megjelenített név",
        "Kedvezmény",
        "Ország",
        "Közterület",
        "Ajtó",
        "Épület",
        "Emelet",
        "Lépcsőház",
        "Házszám",
        "Magánszemély",
        "Magyar illetőség",
        "Adóalanynak nem minősülő gazdasági szervezet",
        "Köz. kívüli adósz.",
        "EU adószám "
    ]

    invoice_numbers = {}
    next_number = 1

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(
            f,
            delimiter=";",
            quoting=csv.QUOTE_MINIMAL
        )

        writer.writerow(headers)

        for row in rows:
            (
                forras_fajl,
                szamla_szam,
                szamla_datum,
                esedekesseg,
                termeknev,
                cikkszam,
                mennyiseg,
                szallito_orszaga,
                gyarto,
                tetel_osszeg,
                valuta,
                brutto_suly,
                brutto_tomeg,
            ) = row

            if szamla_szam not in invoice_numbers:
                invoice_numbers[szamla_szam] = next_number
                next_number += 1

            qty = hu_to_float(mennyiseg) or 0
            amount = hu_to_float(tetel_osszeg) or 0

            # A PDF-ben a tétel teljes összege szerepel,
            # a Kulcs viszont nettó egységárat vár.
            unit_price = amount / qty if qty else 0

            writer.writerow([
                invoice_numbers[szamla_szam],       # Sorszám
                KULCS_PENZNEM,                      # Pénznem
                kulcs_decimal(KULCS_ARFOLYAM),      # Árfolyam
                KULCS_PARTNER_KOD,
                KULCS_PARTNER_NEV,
                KULCS_IRANYITOSZAM,
                KULCS_VAROS,
                KULCS_UTCA,
                "",                                 # Adószám
                KULCS_FIZETESI_MOD,
                kulcs_date(szamla_datum),           # Kelt
                kulcs_date(szamla_datum),           # Teljesítés
                kulcs_date(esedekesseg),  # Esedékesség
                termeknev,
                kulcs_decimal(qty),
                KULCS_MENNYISEGI_EGYSEG,
                kulcs_decimal(unit_price),
                "",
                KULCS_AFA_KULCS,
                "",
                "",
                "",
                cikkszam,                           # TERMEKKOD
                szamla_szam,                        # Bizonylatszám
                "",
                "",
                "",
                "",
                "Lengyelország",
                KULCS_KOZTERULET,
                "",
                "",
                "",
                "",
                KULCS_HAZSZAM,
                0,
                0,                                  # külföldi partner
                0,
                "",
                ""
            ])

# -------------------------
# Fő feldolgozás
# -------------------------
def main():
    input_path = Path(INPUT_DIR)
    output_path = Path(OUTPUT_DIR)
    bad_output_path = Path(BAD_OUTPUT_DIR)

    output_path.mkdir(exist_ok=True)
    bad_output_path.mkdir(exist_ok=True)
    Path(TXT_DEBUG_DIR).mkdir(exist_ok=True)

    if not input_path.exists():
        print(f"Hiba: nem létezik az input mappa: {INPUT_DIR}")
        return

    pdf_files = sorted(input_path.glob("*.pdf"))

    if not pdf_files:
        print("Nem találtam PDF fájlokat az input mappában.")
        return

    print("Google hitelesítés...")
    service = authenticate()

    check_rows = []
    all_output_rows = []

    ok_count = 0
    bad_count = 0

    for pdf_file in pdf_files:
        print(f"Feldolgozás: {pdf_file.name}")

        try:
            text = get_pdf_text(str(pdf_file))

            szamla_szam = parse_invoice_no(text)
            szamla_datum = parse_invoice_date(text)
            esedekesseg = parse_due_date(text)
            data = parse_text(text)

            output_rows = []
            for row in data:
                output_rows.append((
                    pdf_file.name,
                    szamla_szam,
                    szamla_datum,
                    esedekesseg,
                    row[0],
                    row[1],
                    row[2],
                    row[3],
                    row[4],
                    row[5],
                    row[6],
                    row[7],
                    row[8],
                ))

            # Ellenőrzés
            pdf_qty, pdf_amount, pdf_brutto = parse_total_row(text)

            extracted_qty = sum(hu_to_float(r[2]) or 0 for r in data)
            extracted_amount = sum(hu_to_float(r[5]) or 0 for r in data)
            extracted_brutto = sum(hu_to_float(r[8]) or 0 for r in data)

            qty_ok = close_enough(pdf_qty, extracted_qty)
            amount_ok = close_enough(pdf_amount, extracted_amount)
            brutto_ok = close_enough(pdf_brutto, extracted_brutto)

            all_ok = qty_ok and amount_ok and brutto_ok

            excel_name = pdf_file.stem + ".xlsx"

            if all_ok:
                excel_path = output_path / excel_name
                ok_count += 1
                statusz = "OK"
                megjegyzes = ""
                all_output_rows.extend(output_rows)
            else:
                excel_path = bad_output_path / excel_name
                bad_count += 1
                statusz = "HIBAS"
                megjegyzes = "Eltérés az ellenőrzésben"

                # Csak hibás fájlnál mentjük a TXT debugot
                save_txt_debug(pdf_file.name, text)

            save_excel(output_rows, excel_path)

            check_rows.append((
                pdf_file.name,
                szamla_szam,
                szamla_datum,
                len(output_rows),

                pdf_qty,
                extracted_qty,
                "IGEN" if qty_ok else "NEM",

                pdf_amount,
                extracted_amount,
                "IGEN" if amount_ok else "NEM",

                pdf_brutto,
                extracted_brutto,
                "IGEN" if brutto_ok else "NEM",

                statusz,
                megjegyzes,
            ))

            print(f"  -> kinyert sorok: {len(output_rows)}")
            print(f"  -> státusz: {statusz}")
            print(f"  -> mentve: {excel_path}")

        except Exception as e:
            bad_count += 1
            print(f"  HIBA: {pdf_file.name} -> {e}")

            check_rows.append((
                pdf_file.name,
                "",
                "",
                0,

                None,
                None,
                "HIBA",

                None,
                None,
                "HIBA",

                None,
                None,
                "HIBA",

                "HIBA",
                str(e),
            ))

    save_kulcs_csv(all_output_rows, KULCS_OUTPUT_FILE)
    save_check_excel(check_rows, CHECK_FILE)

    print()
    print("KÉSZ")
    print(f"Hibátlan fájlok száma: {ok_count}")
    print(f"Hibás fájlok száma: {bad_count}")
    print(f"Hibátlan Excel mappa: {OUTPUT_DIR}")
    print(f"Hibás Excel mappa: {BAD_OUTPUT_DIR}")
    print(f"Hibás TXT debug mappa: {TXT_DEBUG_DIR}")
    print(f"Ellenőrző fájl: {CHECK_FILE}")
    print(f"Kulcs-Ügyvitel importfájl: {KULCS_OUTPUT_FILE}")


if __name__ == "__main__":
    main()