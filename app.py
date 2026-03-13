import os
import uuid

from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, HTMLResponse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from parser import parse_text
from drive_utils import (
    authenticate,
    pdf_to_google_doc,
    get_doc_text,
    delete_file,
)

app = FastAPI()

TEMP_DIR = "temp"
os.makedirs(TEMP_DIR, exist_ok=True)


@app.get("/", response_class=HTMLResponse)
def home():
    html = """
<!doctype html>
<html lang="hu">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>PDF számla feldolgozó → Excel</title>
  <style>
    :root { color-scheme: dark; }
    body {
      margin: 0;
      font-family: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial;
      background: #0b1020;
      color: #e6e8ee;
    }
    .wrap { max-width: 820px; margin: 0 auto; padding: 28px 18px; }
    .card {
      background: rgba(255,255,255,0.06);
      border: 1px solid rgba(255,255,255,0.12);
      border-radius: 16px;
      padding: 18px;
      box-shadow: 0 12px 40px rgba(0,0,0,0.35);
    }
    h1 { font-size: 20px; margin: 0 0 10px; }
    p { margin: 8px 0; line-height: 1.5; color: rgba(230,232,238,0.9); }
    .row { display: flex; gap: 12px; flex-wrap: wrap; align-items: center; margin-top: 14px; }
    .file {
      flex: 1;
      min-width: 240px;
      padding: 10px 12px;
      background: rgba(0,0,0,0.25);
      border: 1px dashed rgba(255,255,255,0.18);
      border-radius: 12px;
    }
    input[type="file"] { width: 100%; }
    button {
      border: 0;
      border-radius: 12px;
      padding: 10px 14px;
      background: #6d5efc;
      color: white;
      font-weight: 700;
      cursor: pointer;
    }
    button:disabled { opacity: 0.6; cursor: not-allowed; }
    .status { margin-top: 12px; padding: 10px 12px; border-radius: 12px; display:none; }
    .status.ok { display:block; background: rgba(45, 212, 191, 0.15); border: 1px solid rgba(45, 212, 191, 0.35); }
    .status.err { display:block; background: rgba(248, 113, 113, 0.12); border: 1px solid rgba(248, 113, 113, 0.35); }
    .spinner {
      width: 14px; height: 14px;
      border: 2px solid rgba(255,255,255,0.25);
      border-top-color: rgba(255,255,255,0.95);
      border-radius: 50%;
      display:inline-block;
      animation: spin 0.8s linear infinite;
      vertical-align: -2px;
      margin-right: 8px;
    }
    @keyframes spin { to { transform: rotate(360deg); } }
    .footer { margin-top: 10px; opacity: 0.7; font-size: 12px; }
  </style>
</head>
<body>
  <div class="wrap">
    <div class="card">
      <h1>PDF számla feldolgozó → Excel</h1>
      <p>
        Tölts fel egy PDF számlát, és a rendszer elkészíti belőle az Excel fájlt.
      </p>

      <div class="row">
        <div class="file">
          <input id="pdf" type="file" accept="application/pdf" />
          <div class="footer">Kérlek PDF fájlt tölts fel.</div>
        </div>
        <button id="btn">Feldolgozás & letöltés</button>
      </div>

      <div id="status" class="status"></div>
    </div>
  </div>

<script>
  const input = document.getElementById("pdf");
  const btn = document.getElementById("btn");
  const statusBox = document.getElementById("status");

  function setStatus(type, text) {
    statusBox.className = "status " + type;
    statusBox.style.display = "block";
    statusBox.innerHTML = text;
  }

  function clearStatus() {
    statusBox.className = "status";
    statusBox.style.display = "none";
    statusBox.innerHTML = "";
  }

  btn.addEventListener("click", async () => {
    clearStatus();

    const file = input.files?.[0];
    if (!file) {
      setStatus("err", "Kérlek válassz ki egy PDF fájlt.");
      return;
    }

    if (file.type !== "application/pdf") {
      setStatus("err", "Ez nem PDF fájlnak tűnik. Kérlek PDF-et tölts fel.");
      return;
    }

    btn.disabled = true;
    setStatus("ok", `<span class="spinner"></span>Feldolgozás folyamatban...`);

    try {
      const fd = new FormData();
      fd.append("file", file);

      const res = await fetch("/parse", {
        method: "POST",
        body: fd
      });

      if (!res.ok) {
        let msg = "Hiba történt a feldolgozás közben.";
        try {
          const j = await res.json();
          if (j?.detail) {
            msg = typeof j.detail === "string" ? j.detail : JSON.stringify(j.detail);
          }
        } catch {}
        throw new Error(msg);
      }

      const blob = await res.blob();
      const url = URL.createObjectURL(blob);

      const a = document.createElement("a");
      a.href = url;
      a.download = "adatok.xlsx";
      document.body.appendChild(a);
      a.click();
      a.remove();

      URL.revokeObjectURL(url);

      setStatus("ok", "Kész! Az Excel letöltése elindult. ✅");

      setTimeout(() => {
        input.value = "";
        clearStatus();
      }, 1200);

    } catch (e) {
      setStatus("err", "❌ " + (e?.message || "Ismeretlen hiba."));
    } finally {
      btn.disabled = false;
    }
  });
</script>
</body>
</html>
"""
    return HTMLResponse(html)


def parse_hu_number(value):
    """
    Magyar formátumú szám szövegből float:
    '12 345,67' -> 12345.67
    '0,45' -> 0.45
    """
    if value is None:
        return None

    s = str(value).strip()
    if not s:
        return None

    s = s.replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return value


def format_worksheet(ws):
    """
    Excel formázás:
    - sötét fejléc
    - fehér, vastag fejlécszöveg
    - vastagabb keretek
    - váltakozó sorszínek
    - megnevezés oszlop fix szélesség + sortörés
    - többi oszlop automatikus szélesség
    - szűrő és fejléc rögzítés
    """

    # Színek
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    odd_row_fill = PatternFill(fill_type="solid", fgColor="EAF3F8")
    even_row_fill = PatternFill(fill_type="solid", fgColor="D6EAF4")

    header_font = Font(bold=True, color="FFFFFF")
    normal_font = Font(color="000000")

    thick_side = Side(border_style="medium", color="7F7F7F")
    border = Border(
        left=thick_side,
        right=thick_side,
        top=thick_side,
        bottom=thick_side
    )

    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_alignment = Alignment(vertical="top")
    wrap_top_alignment = Alignment(vertical="top", wrap_text=True)

    # Fejléc formázása
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Adatsorok formázása + váltakozó színek
    for row_idx in range(2, ws.max_row + 1):
        row_fill = odd_row_fill if row_idx % 2 == 0 else even_row_fill

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = normal_font
            cell.fill = row_fill
            cell.border = border

            if col_idx == 1:
                cell.alignment = wrap_top_alignment
            else:
                cell.alignment = top_alignment

    # Megnevezés oszlop fix szélesség
    ws.column_dimensions["A"].width = 45

    # Többi oszlop automatikus szélesség
    for col_idx in range(2, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        for cell in ws[col_letter]:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        adjusted_width = min(max_length + 2, 22)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Sorok magassága
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 32

    ws.row_dimensions[1].height = 24

    # Fejléc rögzítése
    ws.freeze_panes = "A2"

    # Szűrő
    ws.auto_filter.ref = ws.dimensions


@app.post("/parse")
async def parse_pdf(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Csak PDF fájl tölthető fel.")

    safe_name = file.filename.replace("/", "_").replace("\\", "_")
    pdf_path = os.path.join(TEMP_DIR, f"{uuid.uuid4()}_{safe_name}")

    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Üres fájl érkezett.")

        with open(pdf_path, "wb") as f:
            f.write(content)

        service = authenticate()
        doc_id = pdf_to_google_doc(service, pdf_path, doc_name=safe_name)

        try:
            text = get_doc_text(service, doc_id)
            data = parse_text(text)

            output_file = os.path.join(TEMP_DIR, f"output_{uuid.uuid4()}.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "Termékek"

            ws.append([
                "Terméknév",
                "Cikkszám",
                "Mennyiség",
                "Szállító országa",
                "Gyártó",
                "Nettó ár",
                "Valuta",
                "Bruttó súly",
                "Bruttó tömeg"
            ])

            for row in data:
                row = list(row)

                # Számoszlopok konvertálása
                # indexek:
                # 0 Terméknév
                # 1 Cikkszám
                # 2 Mennyiség
                # 3 Szállító országa
                # 4 Gyártó
                # 5 Nettó ár
                # 6 Valuta
                # 7 Bruttó súly
                # 8 Bruttó tömeg
                row[2] = parse_hu_number(row[2])
                row[5] = parse_hu_number(row[5])
                row[7] = parse_hu_number(row[7])
                row[8] = parse_hu_number(row[8])

                ws.append(row)

            # Számformátumok
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=3).number_format = '0.00'
                ws.cell(row=row_idx, column=6).number_format = '#,##0.00'
                ws.cell(row=row_idx, column=8).number_format = '0.00'
                ws.cell(row=row_idx, column=9).number_format = '0.00'

            format_worksheet(ws)

            wb.save(output_file)

            return FileResponse(
                output_file,
                filename="adatok.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        finally:
            try:
                delete_file(service, doc_id)
            except Exception:
                pass

    finally:
        try:
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
        except Exception:
            pass
